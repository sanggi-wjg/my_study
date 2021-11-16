import multiprocessing

from openpyxl import load_workbook


class BalanceAccountReader(object):

    def __init__(self, filepath):
        workbook = load_workbook(filepath, read_only = True, data_only = True)
        self._sheet = workbook.worksheets[1]

    @property
    def sheet(self):
        return self._sheet

    def read(self):
        row_limit = 1000
        max_row = self.sheet.max_row
        process_count = (max_row // row_limit) + 1

        results = dict()
        pool = multiprocessing.Pool(processes = process_count)

        for i in range(2, max_row, row_limit):
            start, end = i, i + row_limit
            if end > max_row:
                end = max_row
            res = pool.apply_async(self.read_rows, (start, end))
            results.update(res.get())

        pool.close()
        pool.join()

        return results

    def read_rows(self, start, end):
        read_data = { row[2].value: {
            'invoiceNo': row[2].value,
            'where'    : row[4].value,
            'weight'   : row[5].value,
            'cost'     : row[6].value,
        } for row in self.sheet.iter_rows(min_row = start, max_row = end) }
        return read_data
        # return (start, end)


if __name__ == '__main__':
    reader = BalanceAccountReader('YTO-2021-10-SHA.xlsx')
    result = reader.read()
