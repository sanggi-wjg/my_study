import collections
import sys

from openpyxl import load_workbook


def load(path):
    workbook = load_workbook(path, read_only = True, data_only = True)
    return workbook


def read_current():
    workbook = load('YTO-2021-10-SHA.xlsx')
    sheet = workbook.worksheets[1]

    read_data, duplicate_data = dict(), list()

    for rowNo, row in enumerate(sheet.iter_rows()):
        if rowNo == 0:
            continue
        invoiceNo, where, weight, cost = '', '', '', ''

        for cellNo, cell in enumerate(row):
            if cellNo == 2:
                invoiceNo = str(cell.value).strip()
            elif cellNo == 4:
                where = str(cell.value)
            elif cellNo == 5:
                weight = str(cell.value)
            elif cellNo == 6:
                cost = str(cell.value)

        if not invoiceNo:
            continue

        if read_data.get(invoiceNo, None) is not None:
            duplicate_data.append({
                'invoiceNo': invoiceNo,
                'where'    : where,
                'weight'   : weight,
                'cost'     : cost,
            })
            continue
        else:
            read_data.setdefault(invoiceNo, {
                'invoiceNo': invoiceNo,
                'where'    : where,
                'weight'   : weight,
                'cost'     : cost,
            })
    return read_data, duplicate_data


def read_diff_way_dict():
    workbook = load('YTO-2021-10-SHA.xlsx')
    sheet = workbook.worksheets[1]

    read_data = { row[2].value: {
        'invoiceNo': row[2].value,
        'where'    : row[4].value,
        'weight'   : row[5].value,
        'cost'     : row[6].value,
    } for row in sheet.iter_rows(min_row = 2) }
    return read_data


def read_diff_way_generator_tuple():
    workbook = load('YTO-2021-10-SHA.xlsx')
    sheet = workbook.worksheets[1]

    read_data = (
        (row[2].value, row[2].value, row[4].value, row[5].value, row[6].value,)
        for row in sheet.iter_rows(min_row = 2)
    )
    return read_data


def read_diff_way_deque_tuple():
    workbook = load('YTO-2021-10-SHA.xlsx')
    sheet = workbook.worksheets[1]

    queue = collections.deque(
        ((row[2].value, row[2].value, row[4].value, row[5].value, row[6].value)
         for row in sheet.iter_rows(min_row = 2))
    )
    return queue


def find_current(needle):
    d, _ = read_current()
    print(sys.getsizeof(d))
    print(d[needle])


def find_diff_way_dict(needle):
    d1 = read_diff_way_dict()
    print(sys.getsizeof(d1))
    print(d1[needle])


def find_diff_way_generator_tuple(needle):
    d2 = read_diff_way_generator_tuple()
    print(sys.getsizeof(d2))

    for item in d2:
        if item[0] == needle:
            print(item)
            break


def find_diff_way_deque_tuple(needle):
    d3 = read_diff_way_deque_tuple()

    for item in d3:
        if item[0] == needle:
            print(item)
            break


if __name__ == '__main__':
    target = 'YTG003192008705'
    # find_current(target)

    # find_diff_way_dict(target)

    # find_diff_way_generator_tuple(target)

    find_diff_way_deque_tuple(target)

    a = (1, 2)
    del a[1]
